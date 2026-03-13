from __future__ import annotations
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from decimal import Decimal, InvalidOperation
from datetime import datetime

from inventario import Inventario, Articulo, HEADERS

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

APP_TITLE = "Inventario Refaccionaria BMW"
DB_PATH = "inventario.db"

MONEY_COLS = {"costo", "precio", "precio_mayoreo"}
INT_COLS = {"garantia_meses", "disponibilidad", "vida_util_meses"}
LOW_STOCK_THRESHOLD = 3


def money_fmt(x) -> str:
    try:
        d = Decimal(str(x))
        return f"${d:,.2f}"
    except Exception:
        return str(x)


def parse_money(s: str) -> float:
    s = s.strip().replace("$", "").replace(",", "")
    try:
        return float(Decimal(s))
    except InvalidOperation:
        raise ValueError("Monto inválido (usa números).")


def parse_int(s: str) -> int:
    s = s.strip()
    try:
        return int(Decimal(s))
    except InvalidOperation:
        raise ValueError("Número entero inválido.")


class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1500x780")
        self.minsize(1280, 700)

        self.inv = Inventario(DB_PATH)
        self.inv.cargar()

        self.search_var = tk.StringVar()

        self.f_marca = tk.StringVar(value="Todas")
        self.f_tipo = tk.StringVar(value="Todos")
        self.f_calidad = tk.StringVar(value="Todas")
        self.f_stock = tk.StringVar(value="Todos")

        self.kpi_total = tk.StringVar()
        self.kpi_stock = tk.StringVar()
        self.kpi_cost = tk.StringVar()
        self.kpi_sale = tk.StringVar()
        self.kpi_margin = tk.StringVar()

        self.detail_title = tk.StringVar(value="Selecciona un artículo")
        self.detail_text: tk.Text | None = None
        self.logs_box: tk.Text | None = None

        self.sort_state = {}
        self.status_var = tk.StringVar(value="Listo")

        self._style()
        self._build_ui()
        self._refresh_filters()
        self._render()

        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _on_close(self) -> None:
        try:
            self.inv.close()
        finally:
            self.destroy()

    def _style(self) -> None:
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass

        self.c_bg = "#F6F7FB"
        self.c_card = "#FFFFFF"
        self.c_border = "#E3E6EF"
        self.c_text = "#111827"
        self.c_muted = "#6B7280"
        self.c_accent = "#2563EB"
        self.c_bad = "#DC2626"

        self.configure(bg=self.c_bg)

        self.font_title = ("Segoe UI", 18, "bold")
        self.font_sub = ("Segoe UI", 10)
        self.font_btn = ("Segoe UI", 10, "bold")
        self.font_kpi = ("Segoe UI", 15, "bold")

        style.configure("App.TButton", font=self.font_btn, padding=(12, 10), relief="flat", borderwidth=0)
        style.configure("Primary.App.TButton", font=self.font_btn, padding=(12, 10), background=self.c_accent, foreground="white")
        style.map("Primary.App.TButton", background=[("active", "#1D4ED8")])
        style.configure("Danger.App.TButton", font=self.font_btn, padding=(12, 10), background=self.c_bad, foreground="white")
        style.map("Danger.App.TButton", background=[("active", "#B91C1C")])

        style.configure("App.TEntry", padding=(10, 8), relief="solid")
        style.configure("App.TCombobox", padding=(6, 6))

        style.configure(
            "App.Treeview",
            font=("Segoe UI", 10),
            rowheight=30,
            background=self.c_card,
            fieldbackground=self.c_card,
            foreground=self.c_text,
            borderwidth=0,
        )
        style.configure(
            "App.Treeview.Heading",
            font=("Segoe UI", 10, "bold"),
            padding=(10, 8),
            background="#EEF2FF",
            foreground=self.c_text,
            relief="flat",
        )

    def _build_ui(self) -> None:
        top = tk.Frame(self, bg=self.c_bg)
        top.pack(fill="x", padx=14, pady=(14, 10))

        card = tk.Frame(top, bg=self.c_card, highlightbackground=self.c_border, highlightthickness=1)
        card.pack(fill="x")

        left = tk.Frame(card, bg=self.c_card)
        left.pack(side="left", padx=14, pady=12)
        tk.Label(left, text="Inventario", bg=self.c_card, fg=self.c_text, font=self.font_title).pack(anchor="w")
        tk.Label(left, text="SQLite • Reportes (Excel/PDF) • Auditoría • Backups", bg=self.c_card, fg=self.c_muted, font=self.font_sub).pack(anchor="w", pady=(4, 0))

        right = tk.Frame(card, bg=self.c_card)
        right.pack(side="right", padx=14, pady=12)
        tk.Label(right, text="Buscar", bg=self.c_card, fg=self.c_muted, font=self.font_sub).pack(anchor="e")
        e = ttk.Entry(right, textvariable=self.search_var, width=36, style="App.TEntry")
        e.pack(anchor="e", pady=(4, 0))
        e.bind("<KeyRelease>", lambda _ev: self._render())

        kpi_row = tk.Frame(self, bg=self.c_bg)
        kpi_row.pack(fill="x", padx=14, pady=(0, 10))

        for label, var in [
            ("Artículos", self.kpi_total),
            ("Stock total", self.kpi_stock),
            ("Valor a costo", self.kpi_cost),
            ("Valor a venta", self.kpi_sale),
            ("Margen est.", self.kpi_margin),
        ]:
            box = tk.Frame(kpi_row, bg=self.c_card, highlightbackground=self.c_border, highlightthickness=1)
            box.pack(side="left", expand=True, fill="x", padx=6)
            tk.Label(box, text=label, bg=self.c_card, fg=self.c_muted, font=self.font_sub).pack(anchor="w", padx=12, pady=(10, 0))
            tk.Label(box, textvariable=var, bg=self.c_card, fg=self.c_text, font=self.font_kpi).pack(anchor="w", padx=12, pady=(2, 10))

        bar = tk.Frame(self, bg=self.c_bg)
        bar.pack(fill="x", padx=14, pady=(0, 10))

        fcard = tk.Frame(bar, bg=self.c_card, highlightbackground=self.c_border, highlightthickness=1)
        fcard.pack(side="left", fill="x", expand=True)

        fc = tk.Frame(fcard, bg=self.c_card)
        fc.pack(fill="x", padx=12, pady=10)

        tk.Label(fc, text="Marca", bg=self.c_card, fg=self.c_muted, font=self.font_sub).grid(row=0, column=0, sticky="w")
        self.cb_marca = ttk.Combobox(fc, textvariable=self.f_marca, width=18, state="readonly", style="App.TCombobox")
        self.cb_marca.grid(row=1, column=0, padx=(0, 10))

        tk.Label(fc, text="Tipo", bg=self.c_card, fg=self.c_muted, font=self.font_sub).grid(row=0, column=1, sticky="w")
        self.cb_tipo = ttk.Combobox(fc, textvariable=self.f_tipo, width=18, state="readonly", style="App.TCombobox")
        self.cb_tipo.grid(row=1, column=1, padx=(0, 10))

        tk.Label(fc, text="Calidad", bg=self.c_card, fg=self.c_muted, font=self.font_sub).grid(row=0, column=2, sticky="w")
        self.cb_calidad = ttk.Combobox(fc, textvariable=self.f_calidad, width=18, state="readonly", style="App.TCombobox")
        self.cb_calidad.grid(row=1, column=2, padx=(0, 10))

        tk.Label(fc, text="Stock", bg=self.c_card, fg=self.c_muted, font=self.font_sub).grid(row=0, column=3, sticky="w")
        self.cb_stock = ttk.Combobox(
            fc,
            textvariable=self.f_stock,
            values=["Todos", "Sin stock", "Bajo", "Con stock"],
            width=14,
            state="readonly",
            style="App.TCombobox",
        )
        self.cb_stock.grid(row=1, column=3, padx=(0, 10))

        ttk.Button(fc, text="Limpiar filtros", style="App.TButton", command=self._clear_filters).grid(row=1, column=4, sticky="e")

        for cb in (self.cb_marca, self.cb_tipo, self.cb_calidad, self.cb_stock):
            cb.bind("<<ComboboxSelected>>", lambda _e: self._render())

        acard = tk.Frame(bar, bg=self.c_card, highlightbackground=self.c_border, highlightthickness=1)
        acard.pack(side="right", padx=(10, 0))

        a = tk.Frame(acard, bg=self.c_card)
        a.pack(padx=12, pady=10)
        ttk.Button(a, text="＋ Añadir", style="Primary.App.TButton", command=self._open_add).pack(side="left", padx=(0, 8))
        ttk.Button(a, text="🗑 Eliminar", style="Danger.App.TButton", command=self._delete_selected).pack(side="left", padx=(0, 8))
        ttk.Button(a, text="↻ Recargar", style="App.TButton", command=self._reload).pack(side="left", padx=(0, 8))
        ttk.Button(a, text="Exportar Excel", style="App.TButton", command=self._export_excel).pack(side="left", padx=(0, 8))
        ttk.Button(a, text="Exportar PDF", style="App.TButton", command=self._export_pdf).pack(side="left")

        main = tk.Frame(self, bg=self.c_bg)
        main.pack(fill="both", expand=True, padx=14, pady=(0, 10))

        main.columnconfigure(0, weight=3)
        main.columnconfigure(1, weight=1)
        main.rowconfigure(0, weight=1)

        tcard = tk.Frame(main, bg=self.c_card, highlightbackground=self.c_border, highlightthickness=1)
        tcard.grid(row=0, column=0, sticky="nsew")

        table_frame = tk.Frame(tcard, bg=self.c_card)
        table_frame.pack(fill="both", expand=True, padx=12, pady=12)

        self.tree = ttk.Treeview(table_frame, columns=HEADERS, show="headings", selectmode="browse", style="App.Treeview")
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscroll=vsb.set, xscroll=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        self.tree.tag_configure("even", background="#FFFFFF")
        self.tree.tag_configure("odd", background="#F8FAFF")
        self.tree.tag_configure("stock0", background="#FEE2E2")
        self.tree.tag_configure("stocklow", background="#FEF3C7")

        nice = {
            "id": "ID",
            "nombre_producto": "Producto",
            "costo": "Costo",
            "precio": "Precio",
            "precio_mayoreo": "Mayoreo",
            "tipo_producto": "Tipo",
            "marca": "Marca",
            "calidad": "Calidad",
            "garantia_meses": "Garantía (meses)",
            "disponibilidad": "Stock",
            "vida_util_meses": "Vida útil (meses)",
        }
        for col in HEADERS:
            self.tree.heading(col, text=nice.get(col, col), command=lambda c=col: self._sort_by(c))
            self.tree.column(col, width=120, minwidth=90, anchor="center")

        self.tree.column("nombre_producto", width=320, anchor="w")
        self.tree.column("tipo_producto", width=160, anchor="w")
        self.tree.column("marca", width=120, anchor="w")
        self.tree.column("calidad", width=120, anchor="w")

        self.tree.bind("<<TreeviewSelect>>", lambda _e: self._update_detail())
        self.tree.bind("<Double-1>", self._begin_inline_edit)

        dcard = tk.Frame(main, bg=self.c_card, highlightbackground=self.c_border, highlightthickness=1, width=420)
        dcard.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        dcard.grid_propagate(False)

        d = tk.Frame(dcard, bg=self.c_card)
        d.pack(fill="both", expand=True, padx=12, pady=12)

        tk.Label(d, text="Detalle del artículo", bg=self.c_card, fg=self.c_text, font=("Segoe UI", 13, "bold")).pack(anchor="w")
        tk.Label(d, textvariable=self.detail_title, bg=self.c_card, fg=self.c_muted, font=self.font_sub).pack(anchor="w", pady=(4, 10))

        detail_frame = tk.Frame(d, bg=self.c_card)
        detail_frame.pack(fill="x", expand=False)

        self.detail_text = tk.Text(detail_frame, wrap="word", height=9, bd=0, bg=self.c_card, fg=self.c_text, font=("Segoe UI", 10))
        self.detail_text.pack(side="left", fill="both", expand=True)

        detail_scroll = ttk.Scrollbar(detail_frame, orient="vertical", command=self.detail_text.yview)
        detail_scroll.pack(side="right", fill="y")
        self.detail_text.configure(yscrollcommand=detail_scroll.set, state="disabled")

        tk.Label(d, text="Últimos cambios (auditoría)", bg=self.c_card, fg=self.c_text, font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(14, 6))

        logs_frame = tk.Frame(d, bg=self.c_card)
        logs_frame.pack(fill="both", expand=True)

        self.logs_box = tk.Text(logs_frame, wrap="word", bd=0, bg=self.c_card, fg=self.c_text, font=("Segoe UI", 10))
        self.logs_box.pack(side="left", fill="both", expand=True)

        logs_scroll = ttk.Scrollbar(logs_frame, orient="vertical", command=self.logs_box.yview)
        logs_scroll.pack(side="right", fill="y")
        self.logs_box.configure(yscrollcommand=logs_scroll.set, state="disabled")

        status = tk.Frame(self, bg=self.c_bg)
        status.pack(fill="x", padx=14, pady=(0, 14))
        tk.Label(status, textvariable=self.status_var, bg=self.c_bg, fg=self.c_muted, font=self.font_sub).pack(side="right")

    def _set_text(self, widget: tk.Text | None, content: str) -> None:
        if widget is None:
            return
        widget.configure(state="normal")
        widget.delete("1.0", "end")
        widget.insert("1.0", content)
        widget.configure(state="disabled")

    def _refresh_filters(self) -> None:
        d = self.inv.distinct_values()
        self.cb_marca["values"] = ["Todas"] + d.get("marca", [])
        self.cb_tipo["values"] = ["Todos"] + d.get("tipo_producto", [])
        self.cb_calidad["values"] = ["Todas"] + d.get("calidad", [])

    def _clear_filters(self) -> None:
        self.f_marca.set("Todas")
        self.f_tipo.set("Todos")
        self.f_calidad.set("Todas")
        self.f_stock.set("Todos")
        self.search_var.set("")
        self._render()

    def _reload(self) -> None:
        self.inv.cargar()
        self._refresh_filters()
        self._render()
        self.status_var.set("Recargado.")

    def _filtered_items(self):
        q = self.search_var.get().strip().lower()
        marca = self.f_marca.get()
        tipo = self.f_tipo.get()
        calidad = self.f_calidad.get()
        stockf = self.f_stock.get()

        out = []
        for it in self.inv.items:
            if marca != "Todas" and it.marca != marca:
                continue
            if tipo != "Todos" and it.tipo_producto != tipo:
                continue
            if calidad != "Todas" and it.calidad != calidad:
                continue

            if stockf == "Sin stock" and it.disponibilidad != 0:
                continue
            if stockf == "Bajo" and not (1 <= it.disponibilidad <= LOW_STOCK_THRESHOLD):
                continue
            if stockf == "Con stock" and it.disponibilidad <= 0:
                continue

            if q:
                hay = " ".join([
                    it.id, it.nombre_producto, it.tipo_producto, it.marca, it.calidad,
                    str(it.costo), str(it.precio), str(it.precio_mayoreo),
                    str(it.garantia_meses), str(it.disponibilidad), str(it.vida_util_meses)
                ]).lower()
                if q not in hay:
                    continue

            out.append(it)
        return out

    def _update_kpis(self, items):
        total = len(items)
        stock_total = sum(max(0, it.disponibilidad) for it in items)
        cost_val = sum(max(0, it.disponibilidad) * float(it.costo) for it in items)
        sale_val = sum(max(0, it.disponibilidad) * float(it.precio) for it in items)
        margin = sale_val - cost_val

        self.kpi_total.set(str(total))
        self.kpi_stock.set(str(stock_total))
        self.kpi_cost.set(money_fmt(cost_val))
        self.kpi_sale.set(money_fmt(sale_val))
        self.kpi_margin.set(money_fmt(margin))

    def _row_values(self, it: Articulo):
        vals = []
        for h in HEADERS:
            v = getattr(it, h)
            vals.append(money_fmt(v) if h in MONEY_COLS else v)
        return vals

    def _tag_for(self, it: Articulo, idx: int) -> str:
        if it.disponibilidad == 0:
            return "stock0"
        if 1 <= it.disponibilidad <= LOW_STOCK_THRESHOLD:
            return "stocklow"
        return "even" if idx % 2 == 0 else "odd"

    def _render(self) -> None:
        items = self._filtered_items()
        self._update_kpis(items)

        self.tree.delete(*self.tree.get_children())
        for idx, it in enumerate(items):
            self.tree.insert("", "end", values=self._row_values(it), tags=(self._tag_for(it, idx),))

        self.status_var.set(f"Mostrando {len(items)} / {len(self.inv.items)} artículos.")
        self._update_detail(clear_if_none=True)

    def _get_selected_id(self):
        sel = self.tree.selection()
        if not sel:
            return None
        vals = self.tree.item(sel[0], "values")
        return str(vals[0]).strip() if vals else None

    def _update_detail(self, clear_if_none: bool = False) -> None:
        item_id = self._get_selected_id()
        if not item_id:
            if clear_if_none:
                self.detail_title.set("Selecciona un artículo")
                self._set_text(self.detail_text, "—")
                self._set_text(self.logs_box, "—")
            return

        it = self.inv.buscar_por_id(item_id)
        if not it:
            return

        unit_margin = float(it.precio) - float(it.costo)
        pct = (unit_margin / float(it.precio) * 100) if float(it.precio) else 0.0
        stock = int(it.disponibilidad)
        estado = "SIN STOCK" if stock == 0 else ("STOCK BAJO" if stock <= LOW_STOCK_THRESHOLD else "OK")

        self.detail_title.set(f"{it.id} • {it.nombre_producto}")
        self._set_text(self.detail_text, (
            f"Tipo: {it.tipo_producto}\n"
            f"Marca: {it.marca} | Calidad: {it.calidad}\n"
            f"Precio: {money_fmt(it.precio)} | Costo: {money_fmt(it.costo)} | Mayoreo: {money_fmt(it.precio_mayoreo)}\n"
            f"Margen unitario: {money_fmt(unit_margin)} ({pct:.1f}%)\n"
            f"Stock: {stock}  → Estado: {estado}\n"
            f"Garantía: {it.garantia_meses} meses | Vida útil: {it.vida_util_meses} meses"
        ))

        logs = self.inv.get_logs(item_id, limit=8)
        if not logs:
            self._set_text(self.logs_box, "Sin cambios registrados.")
        else:
            lines = []
            for r in logs:
                ts = r["ts"]
                act = r["action"]
                field = r.get("field") or "-"
                before = r.get("before") or ""
                after = r.get("after") or ""
                if act == "UPDATE":
                    lines.append(f"• {ts}  UPDATE  {field}: {before} → {after}")
                elif act == "ADD":
                    lines.append(f"• {ts}  ADD     creado")
                else:
                    lines.append(f"• {ts}  DELETE  eliminado")
            self._set_text(self.logs_box, "\n".join(lines))

    def _sort_by(self, col: str) -> None:
        children = list(self.tree.get_children(""))
        if not children:
            return

        asc = not self.sort_state.get(col, True)
        self.sort_state[col] = asc
        idx = HEADERS.index(col)

        def key_fn(iid):
            v = self.tree.item(iid, "values")[idx]
            if col in MONEY_COLS:
                try:
                    return Decimal(str(v).replace("$", "").replace(",", ""))
                except Exception:
                    return Decimal(0)
            if col in INT_COLS:
                try:
                    return int(Decimal(str(v)))
                except Exception:
                    return 0
            return str(v).lower()

        children.sort(key=key_fn, reverse=not asc)
        for new_i, iid in enumerate(children):
            self.tree.move(iid, "", new_i)

        for new_i, iid in enumerate(children):
            vals = self.tree.item(iid, "values")
            try:
                stock = int(Decimal(str(vals[HEADERS.index("disponibilidad")])))
            except Exception:
                stock = 0
            if stock == 0:
                tag = "stock0"
            elif 1 <= stock <= LOW_STOCK_THRESHOLD:
                tag = "stocklow"
            else:
                tag = "even" if new_i % 2 == 0 else "odd"
            self.tree.item(iid, tags=(tag,))

        self.status_var.set(f"Ordenado por {col} ({'ASC' if asc else 'DESC'}).")

    def _begin_inline_edit(self, event) -> None:
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return

        row_iid = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        if not row_iid or not col_id:
            return

        col_index = int(col_id.replace("#", "")) - 1
        col_name = HEADERS[col_index]
        if col_name == "id":
            return

        bbox = self.tree.bbox(row_iid, col_id)
        if not bbox:
            return
        x, y, w, h = bbox

        old_value = self.tree.item(row_iid, "values")[col_index]
        entry = ttk.Entry(self.tree, style="App.TEntry")
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, str(old_value))
        entry.focus_set()
        entry.select_range(0, tk.END)

        def cancel(_e=None):
            entry.destroy()

        def commit(_e=None):
            new_raw = entry.get().strip()
            try:
                item_id = str(self.tree.item(row_iid, "values")[0]).strip()
                if not item_id:
                    raise ValueError("ID inválido en la fila.")

                if col_name in MONEY_COLS:
                    new_value = parse_money(new_raw)
                    if new_value < 0:
                        raise ValueError("No se permiten montos negativos.")
                elif col_name in INT_COLS:
                    new_value = parse_int(new_raw)
                    if new_value < 0:
                        raise ValueError("No se permiten valores negativos.")
                else:
                    new_value = new_raw.strip()
                    if new_value == "":
                        raise ValueError("El texto no puede estar vacío.")

                art = self.inv.buscar_por_id(item_id)
                if art:
                    costo = float(new_value) if col_name == "costo" else float(art.costo)
                    precio = float(new_value) if col_name == "precio" else float(art.precio)
                    may = float(new_value) if col_name == "precio_mayoreo" else float(art.precio_mayoreo)
                    if precio < costo:
                        raise ValueError("El precio no puede ser menor que el costo.")
                    if may > precio:
                        raise ValueError("El mayoreo no debe ser mayor que el precio.")

                self.inv.actualizar(item_id, col_name, new_value)

                entry.destroy()
                self._render()
                self.status_var.set(f"Actualizado: {col_name} (ID {item_id})")

            except Exception as e:
                messagebox.showerror("Error", str(e))
                entry.focus_set()

        entry.bind("<Escape>", cancel)
        entry.bind("<Return>", commit)
        entry.bind("<FocusOut>", cancel)

    def _delete_selected(self) -> None:
        item_id = self._get_selected_id()
        if not item_id:
            messagebox.showwarning("Atención", "Selecciona un artículo.")
            return
        if not messagebox.askyesno("Confirmar", f"¿Eliminar el artículo ID {item_id}?"):
            return
        try:
            self.inv.eliminar(item_id)
            self._refresh_filters()
            self._render()
            self.status_var.set("Eliminado (backup creado en /backups).")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _open_add(self) -> None:
        DialogAdd(self, self.inv, on_done=lambda: (self._refresh_filters(), self._render()))

    def _export_excel(self) -> None:
        items = self._filtered_items()
        if not items:
            messagebox.showwarning("Atención", "No hay datos para exportar con los filtros actuales.")
            return

        default_name = f"inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")]
        )
        if not path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Inventario"

            headers = ["ID", "Producto", "Costo", "Precio", "Mayoreo", "Tipo", "Marca", "Calidad", "Garantía (meses)", "Stock", "Vida útil (meses)"]
            ws.append(headers)

            for it in items:
                ws.append([
                    it.id, it.nombre_producto, float(it.costo), float(it.precio), float(it.precio_mayoreo),
                    it.tipo_producto, it.marca, it.calidad, int(it.garantia_meses), int(it.disponibilidad), int(it.vida_util_meses)
                ])

            header_fill = PatternFill("solid", fgColor="EEF2FF")
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            money_cols = [3, 4, 5]
            for r in range(2, ws.max_row + 1):
                for c in money_cols:
                    ws.cell(row=r, column=c).number_format = '"$"#,##0.00'

            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

            wb.save(path)
            messagebox.showinfo("Listo", f"Exportado a Excel:\n{path}")

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar Excel:\n{e}")

    def _export_pdf(self) -> None:
        items = self._filtered_items()
        if not items:
            messagebox.showwarning("Atención", "No hay datos para exportar con los filtros actuales.")
            return

        default_name = f"reporte_inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            initialfile=default_name,
            filetypes=[("PDF", "*.pdf")]
        )
        if not path:
            return

        try:
            styles = getSampleStyleSheet()
            doc = SimpleDocTemplate(
                path,
                pagesize=landscape(letter),
                leftMargin=24,
                rightMargin=24,
                topMargin=24,
                bottomMargin=24
            )

            title = Paragraph(f"<b>Reporte de Inventario</b> — {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Title"])
            subtitle = Paragraph(
                f"Filtros: Marca={self.f_marca.get()} | Tipo={self.f_tipo.get()} | Calidad={self.f_calidad.get()} | Stock={self.f_stock.get()} | Buscar='{self.search_var.get().strip()}'",
                styles["Normal"]
            )

            data = [["ID", "Producto", "Costo", "Precio", "Mayoreo", "Tipo", "Marca", "Calidad", "Garantía", "Stock", "Vida útil"]]
            for it in items:
                data.append([
                    it.id,
                    it.nombre_producto,
                    money_fmt(it.costo),
                    money_fmt(it.precio),
                    money_fmt(it.precio_mayoreo),
                    it.tipo_producto,
                    it.marca,
                    it.calidad,
                    str(it.garantia_meses),
                    str(it.disponibilidad),
                    str(it.vida_util_meses),
                ])

            t = Table(data, repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF2FF")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))

            story = [title, Spacer(1, 8), subtitle, Spacer(1, 12), t]
            doc.build(story)
            messagebox.showinfo("Listo", f"Exportado a PDF:\n{path}")

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar PDF:\n{e}")


class DialogAdd(tk.Toplevel):
    def __init__(self, parent: App, inv: Inventario, on_done):
        super().__init__(parent)
        self.parent = parent
        self.inv = inv
        self.on_done = on_done

        self.title("Añadir artículo")
        self.configure(bg=parent.c_bg)
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        card = tk.Frame(self, bg=parent.c_card, highlightbackground=parent.c_border, highlightthickness=1)
        card.pack(padx=14, pady=14)

        head = tk.Frame(card, bg=parent.c_card)
        head.pack(fill="x", padx=14, pady=(12, 6))
        tk.Label(head, text="Nuevo artículo", bg=parent.c_card, fg=parent.c_text, font=("Segoe UI", 14, "bold")).pack(anchor="w")
        tk.Label(head, text="Completa los campos del inventario.", bg=parent.c_card, fg=parent.c_muted, font=parent.font_sub).pack(anchor="w", pady=(4, 0))

        body = tk.Frame(card, bg=parent.c_card)
        body.pack(fill="x", padx=14, pady=10)

        self.vars = {h: tk.StringVar() for h in HEADERS}
        self.vars["marca"].set("BMW")
        self.vars["calidad"].set("OEM")
        self.vars["garantia_meses"].set("0")
        self.vars["disponibilidad"].set("0")
        self.vars["vida_util_meses"].set("0")

        left = tk.Frame(body, bg=parent.c_card)
        right = tk.Frame(body, bg=parent.c_card)
        left.grid(row=0, column=0, padx=(0, 16))
        right.grid(row=0, column=1)

        labels = {
            "id": "ID (único)",
            "nombre_producto": "Producto",
            "tipo_producto": "Tipo",
            "marca": "Marca",
            "calidad": "Calidad",
            "costo": "Costo ($)",
            "precio": "Precio ($)",
            "precio_mayoreo": "Mayoreo ($)",
            "garantia_meses": "Garantía (meses)",
            "disponibilidad": "Stock",
            "vida_util_meses": "Vida útil (meses)",
        }

        def field(frame, r, key):
            tk.Label(frame, text=labels[key], bg=parent.c_card, fg=parent.c_muted, font=parent.font_sub).grid(row=r, column=0, sticky="w", pady=(0, 4))
            e = ttk.Entry(frame, textvariable=self.vars[key], width=36, style="App.TEntry")
            e.grid(row=r + 1, column=0, sticky="w", pady=(0, 10))
            return e

        r = 0
        for k in ["id", "nombre_producto", "tipo_producto", "marca", "calidad"]:
            field(left, r, k)
            r += 2

        gen_row = tk.Frame(left, bg=parent.c_card)
        gen_row.grid(row=r, column=0, sticky="w", pady=(0, 10))
        ttk.Button(gen_row, text="Generar ID", style="App.TButton", command=self._gen_id).pack(side="left")

        r = 0
        for k in ["costo", "precio", "precio_mayoreo", "garantia_meses", "disponibilidad", "vida_util_meses"]:
            field(right, r, k)
            r += 2

        foot = tk.Frame(card, bg=parent.c_card)
        foot.pack(fill="x", padx=14, pady=(6, 12))
        ttk.Button(foot, text="Cancelar", style="App.TButton", command=self.destroy).pack(side="right", padx=(8, 0))
        ttk.Button(foot, text="Agregar", style="Primary.App.TButton", command=self._add).pack(side="right")

    def _gen_id(self) -> None:
        self.vars["id"].set(self.inv.next_id("BMW"))

    def _add(self) -> None:
        try:
            id_ = self.vars["id"].get().strip()
            if not id_:
                raise ValueError("El ID es obligatorio.")
            nombre = self.vars["nombre_producto"].get().strip()
            if not nombre:
                raise ValueError("El nombre del producto es obligatorio.")

            costo = parse_money(self.vars["costo"].get())
            precio = parse_money(self.vars["precio"].get())
            may = parse_money(self.vars["precio_mayoreo"].get())

            if costo < 0 or precio < 0 or may < 0:
                raise ValueError("Montos no pueden ser negativos.")
            if precio < costo:
                raise ValueError("El precio no puede ser menor que el costo.")
            if may > precio:
                raise ValueError("El mayoreo no debe ser mayor que el precio.")

            art = Articulo(
                id=id_,
                nombre_producto=nombre,
                costo=costo,
                precio=precio,
                precio_mayoreo=may,
                tipo_producto=self.vars["tipo_producto"].get().strip(),
                marca=self.vars["marca"].get().strip(),
                calidad=self.vars["calidad"].get().strip(),
                garantia_meses=parse_int(self.vars["garantia_meses"].get()),
                disponibilidad=parse_int(self.vars["disponibilidad"].get()),
                vida_util_meses=parse_int(self.vars["vida_util_meses"].get()),
            )

            if art.garantia_meses < 0 or art.disponibilidad < 0 or art.vida_util_meses < 0:
                raise ValueError("No se permiten valores negativos.")

            self.inv.agregar(art)
            self.on_done()
            messagebox.showinfo("Listo", "Artículo agregado (backup creado en /backups).")
            self.destroy()

        except Exception as e:
            messagebox.showerror("Error", str(e))


def main() -> None:
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()